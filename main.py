from fastapi import FastAPI, Request, Form, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse, Response
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import ezdxf
from ezdxf.math import Vec2, Vec3, Matrix44
from shapely.geometry import Polygon, Point, LineString
import math
import re
import ezdxf
import math
from ezdxf.math import Matrix44, Vec3
from shapely.geometry import Polygon, Point
from shapely.affinity import affine_transform
from database import store_form_data

from helper import (extract_all_entities,entity_to_geometry,extract_block_text_info_for_insert,
                    determine_room_type_by_entities_ultra_fixed,assign_entities_to_rooms,check_text_within_rooms,clean_room_name,
                    analyze_multi_room_entities)

import uuid
import tempfile
import io
SELF_CHECK_GLOBALS = {
    "total_rooms": 0,
    "room_info": []  # List of {unit_name, room_name} dictionaries
}

 

from Room_dimension.room_dimension import (
    extract_room_boundaries,
    extract_block_text_info,
    clean_text_value,
    check_text_within_room
)
from Room_dimension.dwg_logic import save_copy_with_changes

from db import add_history, get_all_history, clear_all_history, init_db, delete_entry, get_file_content


app = FastAPI()

app.mount("/static", StaticFiles(directory="static"), name="static")

templates = Jinja2Templates(directory="templates")


@app.get("/", response_class=HTMLResponse)
async def login_page(request: Request):
    return templates.TemplateResponse("login.html", {"request": request})


@app.get("/history", response_class=HTMLResponse)
async def view_history(request: Request):
    records = get_all_history()
    return templates.TemplateResponse("history.html", {"request": request, "records": records})


@app.post("/login", response_class=HTMLResponse)
async def login(request: Request, username: str = Form(...), password: str = Form(...)):
    if username == "admin" and password == "admin":
        return RedirectResponse(url="/home", status_code=302)
    return templates.TemplateResponse("login.html", {"request": request, "error": "Login failed"})


@app.get("/home", response_class=HTMLResponse)
async def home_page(request: Request):
    history_data = get_all_history()
    return templates.TemplateResponse("home.html", {"request": request, "history": history_data})


@app.get("/clear-history")
async def clear_history():
    clear_all_history()
    return RedirectResponse("/home", status_code=302)

@app.get("/delete-entry/{entry_id}")
async def delete_entry_route(entry_id: int):
    delete_entry(entry_id)
    return RedirectResponse(url="/history", status_code=302)


@app.get("/self-check", response_class=HTMLResponse)
async def self_check(request: Request):
    return templates.TemplateResponse("result.html", {
        "request": request,
        "check_type": "Self Check",
        "filename": None,
        "rooms": None,
        "submitted_rooms": None,
        "matches": None,
        "download_link": None,
        "excel_link": None,
        "client_file_link": None
    })
    
# Add a new endpoint to analyze DXF and return room data
@app.post("/api/analyze-dxf")
async def analyze_dxf_file_api(file: UploadFile = File(...)):
    """
    API endpoint to analyze DXF file and return room information
    """
    if not file.filename.lower().endswith(".dxf"):
        return {"success": False, "error": "Only .dxf files are supported."}
    
    # Read file content
    client_content = await file.read()
    
    # Create temporary file for processing
    with tempfile.NamedTemporaryFile(suffix=".dxf", delete=False) as temp_file:
        temp_file.write(client_content)
        temp_path = temp_file.name
    
    try:
        # Analyze DXF file and store global variables
        print(f"🔄 Analyzing DXF file: {file.filename}")
        analysis_result = analyze_dxf_file(temp_path, is_self_check=True)
        
        if analysis_result:
            print(f"✅ Analysis completed - {analysis_result['summary']['total_rooms']} rooms found")
            global_data = get_self_check_globals()
            
            return {
                "success": True,
                "total_rooms": global_data["total_rooms"],
                "room_info": global_data["room_info"],
                "unit_names": get_all_unit_names(),
                "room_names": get_all_room_names()
            }
        else:
            print("❌ Analysis failed")
            return {"success": False, "error": "Failed to analyze DXF file"}
            
    except Exception as e:
        print(f"❌ Error analyzing DXF: {str(e)}")
        return {"success": False, "error": f"Error analyzing file: {str(e)}"}
    
    finally:
        # Clean up temporary file
        os.unlink(temp_path)

from fastapi import UploadFile, File, Form

from Referece_check.reference import extract_room_dimensions, extract_doors, compare_values, visualize_mismatches

@app.get("/reference-check", response_class=HTMLResponse)
async def reference_check_page(request: Request):
    return templates.TemplateResponse("reference.html", {
        "request": request,
        "room_mismatches": None,
        "door_mismatches": None,
        "download_link": None,
        "excel_link": None,
        "client_file_link": None
    })


@app.post("/reference-check", response_class=HTMLResponse)
async def reference_check_upload(
    request: Request,
    ref_file: UploadFile = File(...),
    client_file: UploadFile = File(...)
):
    try:
        # Check valid file types
        if not ref_file.filename.endswith(".dxf") or not client_file.filename.endswith(".dxf"):
            return templates.TemplateResponse("reference.html", {
                "request": request,
                "error": "Only .dxf files are supported.",
                "room_mismatches": None,
                "door_mismatches": None,
                "download_link": None
            })

        # Read file contents
        ref_content = await ref_file.read()
        client_content = await client_file.read()

        # Create temporary files for processing
        with tempfile.NamedTemporaryFile(suffix=".dxf", delete=False) as ref_temp:
            ref_temp.write(ref_content)
            ref_temp_path = ref_temp.name

        with tempfile.NamedTemporaryFile(suffix=".dxf", delete=False) as client_temp:
            client_temp.write(client_content)
            client_temp_path = client_temp.name

        try:
            # DXF comparison
            import ezdxf
            ref_doc = ezdxf.readfile(ref_temp_path)
            client_doc = ezdxf.readfile(client_temp_path)

            ref_rooms = extract_room_dimensions(ref_doc)
            client_rooms = extract_room_dimensions(client_doc)
            room_mismatches = compare_values(ref_rooms, client_rooms)

            ref_doors = extract_doors(ref_doc)
            client_doors = extract_doors(client_doc)
            door_mismatches = compare_values(ref_doors, client_doors)

            # Save updated DXF with visualization
            with tempfile.NamedTemporaryFile(suffix="_visualized.dxf", delete=False) as updated_temp:
                updated_temp_path = updated_temp.name

            visualize_mismatches(client_temp_path, room_mismatches, door_mismatches, updated_temp_path)
            
            with open(updated_temp_path, 'rb') as f:
                updated_content = f.read()

            # Generate Excel report
            excel_data = pd.DataFrame({
                "Room Mismatches": room_mismatches,
                "Door Mismatches": door_mismatches
            })
            
            excel_buffer = io.BytesIO()
            excel_data.to_excel(excel_buffer, index=False)
            excel_content = excel_buffer.getvalue()

            # Store in database
            history_id = add_history(
                "Reference Check",
                ref_file.filename,
                client_file.filename,
                client_file.filename.replace('.dxf', '_visualized.dxf'),
                f"report_{uuid.uuid4().hex}.xlsx",
                ref_content,
                client_content,
                updated_content,
                excel_content
            )

            return templates.TemplateResponse("reference.html", {
                "request": request,
                "check_type": "Reference Check",
                "filename": client_file.filename,
                "rooms": None,
                "submitted_rooms": None,
                "matches": None,
                "room_mismatches": room_mismatches,
                "door_mismatches": door_mismatches,
                "download_link": f"/download/modified/{history_id}",
                "excel_link": f"/download/excel/{history_id}",
                "client_file_link": f"/download/client/{history_id}"
            })

        finally:
            # Clean up temporary files
            os.unlink(ref_temp_path)
            os.unlink(client_temp_path)
            if 'updated_temp_path' in locals():
                os.unlink(updated_temp_path)

    except Exception as e:
        return templates.TemplateResponse("reference.html", {
            "request": request,
            "error": f"Error: {str(e)}",
            "room_mismatches": None,
            "door_mismatches": None,
            "download_link": None
        })


def feet_inches_to_mm(feet: int, inches: int) -> float:
    return round((feet * 12 + inches) * 25.4, 2)



def match_user_rooms_to_dxf(submitted_rooms, extracted_rooms):
    """
    Match user input rooms with DXF rooms based on:
    1. Room name text match (case insensitive)
    2. Dimension match with 3% tolerance (both orientations)
    Both conditions must be satisfied for a match.
    """
    matched = []
    unmatched = []
    tolerance_percentage = 0.03  # 3% tolerance for dimensions

    for user_room in submitted_rooms:
        is_matched = False
        target_block = user_room.get("block_name", "").strip()
        
        for dxf_room in extracted_rooms:
            # If user specified a block name, only check rooms in that block
            if target_block:
                room_block = dxf_room.get("BlockName", "").strip()
                if room_block.lower() != target_block.lower():
                    continue
            
            # Condition 1: Check room name match in text
            room_name_matches = any(
                user_room["name"].lower() in t["cleaned"].lower()
                for t in dxf_room["texts"]
            )
            
            # Condition 2: Check dimension match with tolerance
            user_width_mm = user_room["width_mm"]
            user_height_mm = user_room["height_mm"]
            dxf_length = dxf_room.get("Length", 0)
            dxf_breadth = dxf_room.get("Breadth", 0)
            
            # Calculate tolerance values
            width_tolerance = user_width_mm * tolerance_percentage
            height_tolerance = user_height_mm * tolerance_percentage
            
            # Check both orientations: width×height and height×width
            dimension_match_option1 = (
                abs(user_width_mm - dxf_length) <= width_tolerance and
                abs(user_height_mm - dxf_breadth) <= height_tolerance
            )
            
            dimension_match_option2 = (
                abs(user_width_mm - dxf_breadth) <= width_tolerance and
                abs(user_height_mm - dxf_length) <= height_tolerance
            )
            
            dimension_match = dimension_match_option1 or dimension_match_option2

            # BOTH conditions must be satisfied: room name AND dimension match
            if room_name_matches and dimension_match:
                matched.append({
                    "user_room": user_room,
                    "matched_room": dxf_room
                })
                is_matched = True
                print(f"✓ MATCHED: {user_room['name']} - Name: {room_name_matches}, Dimensions: {dimension_match}")
                break

        if not is_matched:
            # Add reason for not matching
            reason_parts = []
            if target_block:
                reason_parts.append(f"Block '{target_block}' specified")
            reason_parts.append("Room name or dimensions don't match within 3% tolerance")
            user_room["reason"] = ", ".join(reason_parts)
            unmatched.append(user_room)
            print(f"✗ NOT MATCHED: {user_room['name']} - {user_room['reason']}")

    return matched, unmatched



def mm_to_feet(mm):
    return round(mm / 304.8, 2)

def mm2_to_sqft(area_mm2):
    return round(area_mm2 / 92903.04, 2)

def export_matches_to_excel(matches, unmatched):
    matched_data = []
    unmatched_data = []

    for match in matches:
        user = match["user_room"]
        dxf = match["matched_room"]
        block_name = dxf.get("BlockName") or dxf.get("Block") or "N/A"

        matched_data.append({
            "Block": block_name,
            "Name of room": user["name"],
            "Length from dxf(feet)": mm_to_feet(dxf.get("Length", 0)),
            "breadth from dxf(feet)": mm_to_feet(dxf.get("Breadth", 0)),
            "Input length(feet)": mm_to_feet(user["width_mm"]),
            "Input breadth(feet)": mm_to_feet(user["height_mm"]),
            "Match": "YES",
            "Reason": "Match",
            "area": mm2_to_sqft(dxf.get("Area", 0))
        })

    for user in unmatched:
        unmatched_data.append({
            "Block": user.get("Block", "N/A"),
            "Name of room": user["name"],
            "Length from dxf(feet)": mm_to_feet(user.get("dxf_length", 0)) if user.get("dxf_length") else "",
            "breadth from dxf(feet)": mm_to_feet(user.get("dxf_breadth", 0)) if user.get("dxf_breadth") else "",
            "Input length(feet)": mm_to_feet(user["width_mm"]),
            "Input breadth(feet)": mm_to_feet(user["height_mm"]),
            "Match": "NO",
            "Reason": user.get("reason", "User Input not match"),
            "area": ""
        })

    # Combine into one DataFrame with separation
    rows = []
    rows.append({"Block": "Matched Room"})  # Section Header
    rows.extend(matched_data)
    rows.append({})  # Empty row
    rows.append({"Block": "Not Matched Room"})  # Section Header
    rows.extend(unmatched_data)

    df = pd.DataFrame(rows)

    # Create Excel in memory
    buffer = io.BytesIO()
    df.to_excel(buffer, index=False, engine='openpyxl')
    
    # Apply colors
    buffer.seek(0)
    wb = load_workbook(buffer)
    ws = wb.active

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        match_cell = row[6]  # Match column
        if match_cell.value == "YES":
            for cell in row:
                cell.fill = green_fill
        elif match_cell.value == "NO":
            for cell in row:
                cell.fill = red_fill

    # Save to buffer
    final_buffer = io.BytesIO()
    wb.save(final_buffer)
    final_buffer.seek(0)
    
    return final_buffer.getvalue()

def create_user_input_colored_dxf(original_dxf_path, submitted_rooms, matched_rooms):
    """
    Create a DXF file that colors ONLY the rooms where BOTH:
    1. Room name text matches user input
    2. Dimensions match user input within 3% tolerance
    
    Colors GREEN (3) only when both conditions are satisfied.
    """
    import ezdxf
    import tempfile
    import os
    from shapely.geometry import Point
    
    try:
        # Load the original DXF
        doc = ezdxf.readfile(original_dxf_path)
        print(f"Processing {len(matched_rooms)} matched rooms for coloring")
        
        entities_colored = 0
        
        # Process each matched room
        for match in matched_rooms:
            matched_room = match["matched_room"]
            user_room = match["user_room"]
            block_name = matched_room.get('BlockName', '')
            room_polygon = matched_room.get('Polygon')
            
            print(f"Coloring room: {user_room['name']} in block: {block_name}")
            
            # Determine where to look for entities
            if block_name and block_name in doc.blocks:
                entity_space = doc.blocks[block_name]
                search_space = "block"
            else:
                entity_space = doc.modelspace()
                search_space = "modelspace"
            
            # Color room boundaries (LWPOLYLINE entities)
            for entity in entity_space:
                if entity.dxftype() == 'LWPOLYLINE' and entity.is_closed:
                    # Check if this polyline represents our matched room
                    if is_matching_room_boundary(entity, matched_room, user_room):
                        entity.dxf.color = 3  # GREEN
                        entities_colored += 1
                        print(f"  ✓ Colored room boundary in {search_space}")
            
            # Color text entities within the room
            if room_polygon:
                for entity in entity_space:
                    if entity.dxftype() in ('TEXT', 'MTEXT'):
                        try:
                            # Get text position
                            if entity.dxftype() == 'TEXT':
                                text_pos = entity.dxf.insert
                                text_content = entity.dxf.text
                            else:  # MTEXT
                                text_pos = entity.dxf.insert
                                text_content = entity.plain_text()
                            
                            # Check if text is within the room boundary
                            text_point = Point(text_pos[0], text_pos[1])
                            if room_polygon.contains(text_point):
                                # Check if text matches room name or dimensions
                                if (user_room['name'].lower() in text_content.lower() or
                                    contains_matching_dimensions(text_content, user_room)):
                                    entity.dxf.color = 3  # GREEN
                                    entities_colored += 1
                                    print(f"  ✓ Colored text: {text_content[:30]}...")
                        except Exception as e:
                            print(f"  Error processing text entity: {e}")
                            continue
        
        # Color INSERT entities (block references) for matched blocks
        matched_block_names = set()
        for match in matched_rooms:
            block_name = match["matched_room"].get('BlockName')
            if block_name:
                matched_block_names.add(block_name)
        
        for entity in doc.modelspace().query('INSERT'):
            if entity.dxf.name in matched_block_names:
                entity.dxf.color = 3  # GREEN
                entities_colored += 1
                print(f"  ✓ Colored block reference: {entity.dxf.name}")
        
        print(f"Total entities colored GREEN: {entities_colored}")
        
        # Save to temporary file
        with tempfile.NamedTemporaryFile(suffix="_user_colored.dxf", delete=False) as temp_file:
            temp_path = temp_file.name
        
        doc.saveas(temp_path)
        
        # Read the file content
        with open(temp_path, 'rb') as f:
            colored_content = f.read()
        
        # Clean up temporary file
        os.unlink(temp_path)
        
        print("✓ Successfully created colored DXF with strict name + dimension matching")
        return colored_content
        
    except Exception as e:
        print(f"Error creating colored DXF: {str(e)}")
        import traceback
        traceback.print_exc()
        
        # Return original file content if coloring fails
        try:
            with open(original_dxf_path, 'rb') as f:
                return f.read()
        except:
            return b""
        
        
        
def is_matching_room_boundary(entity, matched_room, user_room):
    """
    Check if the entity boundary matches the room by comparing handle or dimensions
    """
    try:
        # First try handle matching (most reliable)
        room_handle = matched_room.get('Handle')
        if room_handle and hasattr(entity.dxf, 'handle') and entity.dxf.handle == room_handle:
            return True
        
        # Fallback: Compare dimensions if polygon data available
        if hasattr(entity, 'get_points') and matched_room.get('Polygon'):
            points = [p[:2] for p in entity.get_points()]
            if len(points) >= 3:
                from shapely.geometry import Polygon
                try:
                    entity_polygon = Polygon(points)
                    if entity_polygon.is_valid:
                        matched_polygon = matched_room['Polygon']
                        
                        # Compare areas with tolerance
                        area_diff = abs(entity_polygon.area - matched_polygon.area)
                        area_tolerance = matched_polygon.area * 0.05  # 5% tolerance for area
                        
                        if area_diff <= area_tolerance:
                            return True
                except:
                    pass
        
        return False
        
    except Exception as e:
        print(f"Error in boundary matching: {e}")
        return False

def contains_matching_dimensions(text_content, user_room):
    """
    Check if text contains dimensions that match user input within 3% tolerance
    """
    try:
        import re
        
        # Convert user input to millimeters
        user_width_mm = user_room["width_mm"]
        user_height_mm = user_room["height_mm"]
        
        # Calculate 3% tolerance
        width_tolerance = user_width_mm * 0.03
        height_tolerance = user_height_mm * 0.03
        
        # Look for dimension patterns like "12'0\"x13'11\""
        dimension_pattern = r"(\d+)'(\d+)\"x(\d+)'(\d+)\""
        match = re.search(dimension_pattern, text_content)
        
        if match:
            # Extract dimensions from text
            l_ft, l_in, b_ft, b_in = map(int, match.groups())
            
            # Convert to millimeters
            text_l_mm = (l_ft * 12 + l_in) * 25.4
            text_b_mm = (b_ft * 12 + b_in) * 25.4
            
            # Check both orientations
            match_option1 = (
                abs(user_width_mm - text_l_mm) <= width_tolerance and
                abs(user_height_mm - text_b_mm) <= height_tolerance
            )
            
            match_option2 = (
                abs(user_width_mm - text_b_mm) <= width_tolerance and
                abs(user_height_mm - text_l_mm) <= height_tolerance
            )
            
            return match_option1 or match_option2
        
        # Look for individual dimensions like "12'0\"" or "13'11\""
        individual_dims = re.findall(r"(\d+)'(\d+)\"", text_content)
        if individual_dims:
            user_dims_mm = [user_width_mm, user_height_mm]
            user_tolerances = [width_tolerance, height_tolerance]
            
            for dim_ft, dim_in in individual_dims:
                dim_mm = (int(dim_ft) * 12 + int(dim_in)) * 25.4
                
                # Check if this dimension matches either user dimension
                for user_dim, tolerance in zip(user_dims_mm, user_tolerances):
                    if abs(user_dim - dim_mm) <= tolerance:
                        return True
        
        return False
        
    except Exception as e:
        print(f"Error in dimension text matching: {e}")
        return False



def add_user_input_visualization_to_dxf(doc, matched_rooms):
    """
    Add visual indicators for user input rooms with block information
    """
    for i, match in enumerate(matched_rooms):
        user_room = match["user_room"]
        matched_room = match["matched_room"]
        block_name = matched_room.get('BlockName', 'Unknown')
        user_specified_block = user_room.get('block_name', '')
        
        if 'Polygon' in matched_room and matched_room['Polygon']:
            bounds = matched_room['Polygon'].bounds
            
            # Create a highlight rectangle around the matched room
            margin = 200
            rect_points = [
                (bounds[0] - margin, bounds[1] - margin),
                (bounds[2] + margin, bounds[1] - margin),
                (bounds[2] + margin, bounds[3] + margin),
                (bounds[0] - margin, bounds[3] + margin),
                (bounds[0] - margin, bounds[1] - margin)
            ]
            
            # Add highlight rectangle with dashed line
            doc.modelspace().add_lwpolyline(
                rect_points,
                dxfattribs={
                    'color': 2,      # Yellow
                    'linetype': 'DASHED',
                    'lineweight': 100
                }
            )
            
            # Calculate center for labels
            center_x = (bounds[0] + bounds[2]) / 2
            center_y = (bounds[1] + bounds[3]) / 2
            
            # Add "USER MATCHED" label with block info
            room_label = f"✓ USER MATCHED: {user_room['name'].upper()} (Block: {block_name})"
            doc.modelspace().add_text(
                room_label,
                dxfattribs={
                    'height': 100,
                    'insert': (center_x, center_y + 350),
                    'color': 2,  # Yellow
                    'style': 'Standard'
                }
            )
            
            # Add user specified block info if different from actual
            if user_specified_block.lower() != block_name.lower():
                block_info = f"USER SPECIFIED BLOCK: {user_specified_block}"
                doc.modelspace().add_text(
                    block_info,
                    dxfattribs={
                        'height': 70,
                        'insert': (center_x, center_y + 250),
                        'color': 1,  # Red - to indicate mismatch
                        'style': 'Standard'
                    }
                )
            
            # Add user input dimensions
            user_dim_text = f"USER: {user_room['width_feet']}'{user_room['width_inches']}\"x{user_room['height_feet']}'{user_room['height_inches']}\""
            doc.modelspace().add_text(
                user_dim_text,
                dxfattribs={
                    'height': 80,
                    'insert': (center_x, center_y + 150),
                    'color': 1,  # Red
                    'style': 'Standard'
                }
            )
            
            # Add DXF dimensions for comparison
            dxf_dim_text = f"DXF: {matched_room.get('LengthStr', 'N/A')}x{matched_room.get('BreadthStr', 'N/A')}"
            doc.modelspace().add_text(
                dxf_dim_text,
                dxfattribs={
                    'height': 80,
                    'insert': (center_x, center_y),
                    'color': 3,  # Green
                    'style': 'Standard'
                }
            )
            
            # Add area information
            area_sqft = round(matched_room.get('Area', 0) / 92903.04, 2)
            area_text = f"AREA: {area_sqft} sq ft"
            doc.modelspace().add_text(
                area_text,
                dxfattribs={
                    'height': 60,
                    'insert': (center_x, center_y - 150),
                    'color': 6,  # Magenta
                    'style': 'Standard'
                }
            )




# Helper function to import the required functions from your modules
def check_text_within_room(room, text_info):
    """Use your existing function"""
    from shapely.geometry import Point
    return room['Polygon'].contains(Point(text_info['Position']))


def mm_to_feet_inches_label(mm):
    """Use your existing function"""
    inches = mm / 25.4
    feet = int(inches // 12)
    inch_rem = round(inches % 12)
    if inch_rem == 12:
        feet += 1
        inch_rem = 0
    return f"{feet}'{inch_rem}\""

@app.post("/upload-{check_type}", response_class=HTMLResponse)
async def upload_dxf_with_block(
    request: Request,
    check_type: str,
    file: UploadFile = File(...),
    room_count: int = Form(...),
):
    form = await request.form()
    check_type_display = check_type.replace("-", " ").title()

    if not file.filename.lower().endswith(".dxf"):
        return templates.TemplateResponse("result.html", {
            "request": request,
            "check_type": check_type_display,
            "error": "Only .dxf files are supported.",
            "filename": None,
            "rooms": None,
            "submitted_rooms": None,
            "matches": None,
            "download_link": None,
            "excel_link": None,
            "client_file_link": None
        })

    # Read file content
    client_content = await file.read()

    # Create temporary file for processing
    with tempfile.NamedTemporaryFile(suffix=".dxf", delete=False) as temp_file:
        temp_file.write(client_content)
        temp_path = temp_file.name

    try:
        # NEW: Analyze DXF file and store global variables for self-check
        if check_type == "self-check":
            print(f"🔄 Analyzing DXF file for self-check: {file.filename}")
            analysis_result = analyze_dxf_file(temp_path, is_self_check=True)
            
            if analysis_result:
                print(f"✅ Analysis completed - {analysis_result['summary']['total_rooms']} rooms found")
                print(f"📊 Global variables updated: {get_self_check_globals()}")
            else:
                print("❌ Analysis failed")
        
        # Original room extraction logic
        rooms = extract_room_boundaries(temp_path)
        texts = extract_block_text_info(temp_path)

        all_texts = []
        for blk_texts in texts.values():
            for text in blk_texts:
                text["original"] = text["Text"]
                text["cleaned"] = clean_text_value(text["Text"])
                all_texts.append(text)

        for room in rooms:
            room["texts"] = [t for t in all_texts if check_text_within_room(room, t)]

        submitted_rooms = []
        for i in range(1, room_count + 1):
            name = form.get(f"room_name_{i}", "").strip().lower()
            width_ft = int(form.get(f"width_feet_{i}", 0))
            width_in = int(form.get(f"width_inches_{i}", 0))
            height_ft = int(form.get(f"height_feet_{i}", 0))
            height_in = int(form.get(f"height_inches_{i}", 0))
            block_name = form.get(f"block_name_{i}", "").strip()

            submitted_rooms.append({
                "name": name,
                "width_mm": feet_inches_to_mm(width_ft, width_in),
                "height_mm": feet_inches_to_mm(height_ft, height_in),
                "width_feet": width_ft,
                "width_inches": width_in,
                "height_feet": height_ft,
                "height_inches": height_in,
                "block_name": block_name
            })

        # NEW: Store form data in database
        try:
            user_name = form.get("user_name", "Anonymous")  # Get user name from form
            form_submission_id = store_form_data(user_name, file.filename, room_count, submitted_rooms)
            print(f"✅ Form data stored in database with ID: {form_submission_id}")
        except Exception as e:
            print(f"❌ Error storing form data: {str(e)}")
            # Continue processing even if database storage fails

        # Use the modified matching function
        matches, unmatched = match_user_rooms_to_dxf(submitted_rooms, rooms)

        # Generate Excel report in memory
        excel_content = export_matches_to_excel(matches, unmatched)

        # Create updated DXF file (standard update)
        with tempfile.NamedTemporaryFile(suffix="_updated.dxf", delete=False) as updated_temp:
            updated_temp_path = updated_temp.name

        save_copy_with_changes(temp_path, updated_temp_path, rooms, texts)
        
        with open(updated_temp_path, 'rb') as f:
            updated_content = f.read()

        # For self-check, create an additional DXF with block-specific user input coloring
        reference_content = b""
        reference_filename = ""
        
        if check_type == "self-check":
            reference_content = create_user_input_colored_dxf(temp_path, submitted_rooms, matches)
            reference_filename = f"user_input_colored_{file.filename}"
        else:
            reference_filename = form.get("reference_filename", "")

        # Store in database
        history_id = add_history(
            check_type_display,
            reference_filename,
            file.filename,
            f"updated_{file.filename}",
            "full_report.xlsx",
            reference_content,
            client_content,
            updated_content,
            excel_content
        )

        # NEW: Add global check information to template context for self-check
        template_context = {
            "request": request,
            "check_type": check_type_display,
            "filename": file.filename,
            "rooms": rooms,
            "submitted_rooms": submitted_rooms,
            "matches": matches,
            "download_link": f"/download/modified/{history_id}",
            "excel_link": f"/download/excel/{history_id}",
            "client_file_link": f"/download/client/{history_id}",
            "reference_link": f"/download/reference/{history_id}" if check_type == "self-check" else None
        }
        
        # Add global check data for self-check
        if check_type == "self-check":
            global_data = get_self_check_globals()
            template_context.update({
                "global_check_data": global_data,
                "total_rooms_analyzed": global_data["total_rooms"],
                "room_info": global_data["room_info"],
                "unit_names": get_all_unit_names(),
                "all_room_names": get_all_room_names()
            })

        return templates.TemplateResponse("result.html", template_context)

    finally:
        # Clean up temporary files
        os.unlink(temp_path)
        if 'updated_temp_path' in locals():
            os.unlink(updated_temp_path)


 

@app.get("/download/reference/{history_id}")
async def download_reference_file(history_id: int):
    file_content, filename = get_file_content(history_id, "reference")
    if file_content:
        return Response(
            content=file_content,
            media_type="application/octet-stream",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    return HTMLResponse("<h3>File not found</h3>", status_code=404)


@app.get("/download/client/{history_id}")
async def download_client_file(history_id: int):
    file_content, filename = get_file_content(history_id, "client")
    if file_content:
        return Response(
            content=file_content,
            media_type="application/octet-stream",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    return HTMLResponse("<h3>File not found</h3>", status_code=404)


@app.get("/download/modified/{history_id}")
async def download_modified_file(history_id: int):
    file_content, filename = get_file_content(history_id, "modified")
    if file_content:
        return Response(
            content=file_content,
            media_type="application/octet-stream",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    return HTMLResponse("<h3>File not found</h3>", status_code=404)


@app.get("/download/excel/{history_id}")
async def download_excel_file(history_id: int):
    file_content, filename = get_file_content(history_id, "excel")
    if file_content:
        return Response(
            content=file_content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    return HTMLResponse("<h3>File not found</h3>", status_code=404)


# Legacy download endpoint for backward compatibility
@app.get("/download/{filename}")
async def download_file(filename: str):
    return HTMLResponse(f"<h3>File not found: {filename}</h3>", status_code=404)


 



def analyze_dxf_file(file_path, is_self_check=False):
    """
    Analyze DXF file and extract comprehensive room information
    Based on the main() function logic
    """
    global SELF_CHECK_GLOBALS
    
    try:
        # Load DXF file
        doc = ezdxf.readfile(file_path)
        print(f"✓ Loaded: {file_path}")
    except Exception as e:
        print(f"✗ Error loading DXF: {e}")
        return None

    # Initialize data structures
    all_rooms = []
    global_stats = {
        "units_processed": 0,
        "total_rooms": 0,
        "total_entities": 0,
        "assigned_entities": 0,
        "detection_methods": {"auto": 0, "text": 0, "fallback": 0},
        "room_types": {},
        "size_distribution": {"SMALL": 0, "MEDIUM": 0, "LARGE": 0}
    }

    # Process UNIT-1 inserts
    unit_inserts = list(doc.modelspace().query('INSERT[name=="UNIT-1"]'))
    if not unit_inserts:
        print("✗ No UNIT-1 inserts found")
        return None

    # Store all entities globally for door line search
    all_entities_global = []

    for insert_idx, top_insert in enumerate(unit_inserts, 1):
        global_stats["units_processed"] += 1

        # Calculate transformation matrix
        try:
            top_mat = Matrix44.chain(
                Matrix44.translate(*top_insert.dxf.insert),
                Matrix44.z_rotate(math.radians(top_insert.dxf.rotation)),
                Matrix44.scale(top_insert.dxf.xscale, top_insert.dxf.yscale, top_insert.dxf.zscale),
            )
        except Exception as e:
            print(f"✗ Unit {insert_idx}: Matrix calculation failed - {e}")
            continue

        # Extract all entities (assuming this function exists in helper.py)
        insert_entities = extract_all_entities(doc, "UNIT-1", top_mat)
        all_entities_global.extend(insert_entities)
        global_stats["total_entities"] += len(insert_entities)

        # Identify room boundaries
        rooms = []
        room_boundary_handles = set()
        for entity, mat, handle, depth in insert_entities:
            if entity.dxf.layer.strip() == "A-Room Boundary":
                geom = entity_to_geometry(entity, mat)
                if isinstance(geom, Polygon) and geom.is_valid and geom.area > 1:
                    room_data = {
                        "polygon": geom,
                        "entities_inside": [],
                        "unit_id": top_insert.dxf.handle,
                        "unit_num": insert_idx,
                        "unit_name": f"Unit-{insert_idx}",
                        "boundary_handle": handle,
                        "area": geom.area,
                        "depth": depth,
                        "room_name": "Unknown",
                        "auto_detected_name": None,
                        "text_detected_name": None,
                        "detection_method": "none",
                        "size_category": "UNKNOWN",
                        # Add fields for compatibility with existing template structure
                        "texts": [],  # Will be populated later
                        "bounds": {
                            "min_x": geom.bounds[0],
                            "min_y": geom.bounds[1], 
                            "max_x": geom.bounds[2],
                            "max_y": geom.bounds[3]
                        }
                    }
                    rooms.append(room_data)
                    room_boundary_handles.add(handle)
        global_stats["total_rooms"] += len(rooms)

        if not rooms:
            print(f"✗ Unit {insert_idx}: No room boundaries found")
            continue

        # Extract text information (assuming this function exists in helper.py)
        all_block_texts = extract_block_text_info_for_insert(doc, "UNIT-1", top_mat)

        # Assign entities to rooms (assuming this function exists in helper.py)
        assignment_stats, unassigned_entities = assign_entities_to_rooms(
            rooms, insert_entities, room_boundary_handles
        )
        global_stats["assigned_entities"] += assignment_stats["assigned_entities"]

        # Determine room names and types
        for room_idx, room in enumerate(rooms):
            auto_detected_name = None
            text_detected_name = None

            # Auto-detection (assuming this function exists in helper.py)
            auto_detected_name = determine_room_type_by_entities_ultra_fixed(
                room["entities_inside"],
                room["area"],
                rooms,
                doc,
                all_entities_global
            )

            # Text-based detection
            room_texts = [
                text for text in all_block_texts
                if check_text_within_rooms(room["polygon"], text['Position'])
            ]
            
            # Populate texts field for compatibility
            room["texts"] = room_texts

            if room_texts:
                room_labels = [
                    text for text in room_texts
                    if not ('x' in text['Text'].lower() and '"' in text['Text'])
                ]
                if room_labels:
                    room_labels.sort(key=lambda x: len(x['Text']))
                    text_detected_name = clean_room_name(room_labels[0]['Text'])
                else:
                    text_detected_name = clean_room_name(room_texts[0]['Text'])

            # Set room name and detection method
            room["auto_detected_name"] = auto_detected_name
            room["text_detected_name"] = text_detected_name

            if auto_detected_name:
                room["room_name"] = auto_detected_name
                room["detection_method"] = "auto"
                global_stats["detection_methods"]["auto"] += 1
            elif text_detected_name:
                room["room_name"] = text_detected_name
                room["detection_method"] = "text"
                global_stats["detection_methods"]["text"] += 1
            else:
                room["room_name"] = f"Room_{room_idx + 1}"
                room["detection_method"] = "fallback"
                global_stats["detection_methods"]["fallback"] += 1

            room_type = room["room_name"]
            global_stats["room_types"][room_type] = global_stats["room_types"].get(room_type, 0) + 1

        all_rooms.extend(rooms)
        print(f"✓ Unit {insert_idx}: {len(rooms)} rooms processed")

    # Store in global variables ONLY for self-check
    if is_self_check:
        SELF_CHECK_GLOBALS["total_rooms"] = global_stats["total_rooms"]
        SELF_CHECK_GLOBALS["room_info"] = [
            {
                "unit_name": room["unit_name"],
                "room_name": room["auto_detected_name"] if room["auto_detected_name"] else room["room_name"]
            }
            for room in all_rooms
        ]
        print(f"✓ Self-check globals updated: {SELF_CHECK_GLOBALS['total_rooms']} rooms")
        print(f"✓ Room info: {SELF_CHECK_GLOBALS['room_info']}")

    # Calculate size categories
    if all_rooms:
        areas = sorted([room["area"] for room in all_rooms])
        if len(areas) > 2:
            small_threshold = areas[len(areas) // 3]
            large_threshold = areas[2 * len(areas) // 3]
        else:
            small_threshold = areas[0] if areas else 0
            large_threshold = areas[-1] if areas else 0

        for room in all_rooms:
            area = room["area"]
            if area <= small_threshold:
                room["size_category"] = "SMALL"
            elif area <= large_threshold:
                room["size_category"] = "MEDIUM"
            else:
                room["size_category"] = "LARGE"
            global_stats["size_distribution"][room["size_category"]] += 1

    # Global multi-room analysis (assuming this function exists in helper.py)
    global_multi_room_analysis = analyze_multi_room_entities(all_rooms)

    # Prepare summary data for template
    analysis_summary = {
        "total_units": global_stats["units_processed"],
        "total_rooms": global_stats["total_rooms"],
        "total_entities": global_stats["total_entities"],
        "assigned_entities": global_stats["assigned_entities"],
        "assignment_rate": (global_stats["assigned_entities"] / max(global_stats["total_entities"], 1)) * 100,
        "detection_methods": global_stats["detection_methods"],
        "room_types": global_stats["room_types"],
        "size_distribution": global_stats["size_distribution"],
        "multi_room_entities": len(global_multi_room_analysis)
    }

    # Group rooms by unit for better organization
    units_data = {}
    for room in all_rooms:
        unit_num = room["unit_num"]
        unit_name = room["unit_name"]
        if unit_num not in units_data:
            units_data[unit_num] = {
                "unit_name": unit_name,
                "unit_number": unit_num,
                "rooms": []
            }
        units_data[unit_num]["rooms"].append(room)

    # Convert polygon to serializable format for template compatibility
    serializable_rooms = []
    for room in all_rooms:
        room_copy = room.copy()
        # Convert polygon to bounds for template usage
        if hasattr(room_copy["polygon"], "bounds"):
            room_copy["bounds"] = {
                "min_x": room_copy["polygon"].bounds[0],
                "min_y": room_copy["polygon"].bounds[1],
                "max_x": room_copy["polygon"].bounds[2],
                "max_y": room_copy["polygon"].bounds[3]
            }
        # Remove polygon object as it's not JSON serializable
        room_copy.pop("polygon", None)
        
        # Ensure template compatibility - add capitalized attributes
        if "area" in room_copy:
            room_copy["Area"] = room_copy["area"]
        if "room_name" in room_copy:
            room_copy["Name"] = room_copy["room_name"]  # Some templates might expect "Name"
        
        serializable_rooms.append(room_copy)

    return {
        "summary": analysis_summary,
        "units": units_data,
        "all_rooms": serializable_rooms,
        "multi_room_analysis": global_multi_room_analysis,
        "room_count": global_stats["total_rooms"],
        "unit_names": [unit_data["unit_name"] for unit_data in units_data.values()],
        "room_names": [room["room_name"] for room in all_rooms]
    }


def get_self_check_globals():
    """
    Get the current self-check global variables
    """
    return SELF_CHECK_GLOBALS.copy()


def get_room_info_by_unit(unit_name):
    """
    Get all rooms for a specific unit from self-check globals
    """
    return [room for room in SELF_CHECK_GLOBALS["room_info"] if room["unit_name"] == unit_name]


def get_all_room_names():
    """
    Get all room names from self-check globals
    """
    return [room["room_name"] for room in SELF_CHECK_GLOBALS["room_info"]]


def get_all_unit_names():
    """
    Get unique unit names from self-check globals
    """
    return list(set(room["unit_name"] for room in SELF_CHECK_GLOBALS["room_info"]))


def reset_self_check_globals():
    """
    Reset the self-check global variables
    """
    global SELF_CHECK_GLOBALS
    SELF_CHECK_GLOBALS = {
        "total_rooms": 0,
        "room_info": []
    }


init_db()